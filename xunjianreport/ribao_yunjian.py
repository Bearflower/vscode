import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


# 设置起始日期和结束日期
start_date = datetime(2025, 1, 25)
end_date = datetime(2025, 1, 28)

# 模板文件路径
template_path = "/Users/yl/vscode/xunjianreport/report_demo/yj20250124.xlsx" #需要改路径

# 图片文件路径 需要改路径
image1_path = "/Users/yl/vscode/xunjianreport/yunjian_pic/1.png" #需要改路径
image2_path = "/Users/yl/vscode/xunjianreport/yunjian_pic/2.png" #需要改路径

# 输出目录（可根据需要更改）
output_dir = '/Users/yl/vscode/xunjianreport/yjdaily_reports' #需要改路径
os.makedirs(output_dir, exist_ok=True)

# 遍历日期范围内的每一天
current_date = start_date
while current_date <= end_date:
    date_str = current_date.strftime('%Y%m%d')
    
    # 加载模板文件
    wb = load_workbook(template_path)
    
    # 修改 "功能页面" sheet 的 B2:B14 单元格
    if '功能页面' in wb.sheetnames:
        ws_func_page = wb['功能页面']
        for row in range(2, 15):  # B2 到 B14 对应第 2 行到第 14 行
            ws_func_page[f'B{row}'] = current_date.strftime('%Y-%m-%d')

    # 修改 "过程执行情况" sheet 的 A2 单元格
    if '过程执行情况' in wb.sheetnames:
        ws_process_exec = wb['过程执行情况']
        ws_process_exec['A2'] = current_date.strftime('%Y-%m-%d')
    
    # 插入图片到"首页数据"sheet
    if '首页数据' in wb.sheetnames and os.path.exists(image1_path):
        ws_home_data = wb['首页数据']
        img1 = Image(image1_path)
        # 设置图片插入位置（可以根据需要调整）
        ws_home_data.add_image(img1, 'A1')
    
    # 插入图片到"业扩和光伏一张图数据"sheet
    if '业扩和光伏一张图数据' in wb.sheetnames and os.path.exists(image2_path):
        ws_chart_data = wb['业扩和光伏一张图数据']
        img2 = Image(image2_path)
        # 设置图片插入位置（可以根据需要调整）
        ws_chart_data.add_image(img2, 'A1')
        
    # 处理"数据巡检"sheet
    if '数据巡检' in wb.sheetnames:
        ws_data_inspection = wb['数据巡检']
        # 计算从起始日期到当前日期的天数
        days_count = (current_date - start_date).days + 1
        
        # 在CCG列之后插入所需的列数（每天2列）
        ccg_column_index = 2112  # CCG是第2112列
        total_new_columns = days_count * 2
        ws_data_inspection.insert_cols(ccg_column_index + 1, total_new_columns)
        
        # 填充标题行和子标题行，并复制格式
        for i in range(days_count):
            # 计算当前日期
            target_date = start_date + timedelta(days=i)
            date_str_header = target_date.strftime('%Y-%m-%d')
            
            # 每个日期占用两列，起始列索引
            start_col_index = ccg_column_index + 1 + (i * 2)
            
            # 获取参考单元格的格式（从已有列复制）
            reference_cell = ws_data_inspection.cell(row=1, column=ccg_column_index)  # 参考CCG列
            
            # 第一行填写日期并复制格式
            date_cell_1 = ws_data_inspection.cell(row=1, column=start_col_index, value=date_str_header)
            date_cell_2 = ws_data_inspection.cell(row=1, column=start_col_index + 1, value=date_str_header)
            
            # 合并相同日期的两个单元格
            ws_data_inspection.merge_cells(start_row=1, start_column=start_col_index, 
                                         end_row=1, end_column=start_col_index + 1)
            
            # 获取合并后的单元格区域
            merged_cell = ws_data_inspection.cell(row=1, column=start_col_index)
            
            # 设置合并单元格的格式：字体宋体，字号12号，垂直居中，水平居中
            from openpyxl.styles import Font, Alignment
            merged_cell.font = Font(name='宋体', size=12)
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 如果参考单元格有其他样式，则也复制
            if reference_cell.has_style:
                merged_cell.border = reference_cell.border.copy()
                merged_cell.fill = reference_cell.fill.copy()
                merged_cell.number_format = reference_cell.number_format
                merged_cell.protection = reference_cell.protection.copy()
            
            # 第二行填写工单数量和平均时间并复制格式
            header_cell_1 = ws_data_inspection.cell(row=2, column=start_col_index, value="工单数量")
            header_cell_2 = ws_data_inspection.cell(row=2, column=start_col_index + 1, value="平均时间")
            
            # 参考第二行的格式
            ref_header_cell = ws_data_inspection.cell(row=2, column=ccg_column_index)
            if ref_header_cell.has_style:
                header_cell_1.font = ref_header_cell.font.copy()
                header_cell_1.border = ref_header_cell.border.copy()
                header_cell_1.fill = ref_header_cell.fill.copy()
                header_cell_1.number_format = ref_header_cell.number_format
                header_cell_1.protection = ref_header_cell.protection.copy()
                header_cell_1.alignment = ref_header_cell.alignment.copy()
                
                header_cell_2.font = ref_header_cell.font.copy()
                header_cell_2.border = ref_header_cell.border.copy()
                header_cell_2.fill = ref_header_cell.fill.copy()
                header_cell_2.number_format = ref_header_cell.number_format
                header_cell_2.protection = ref_header_cell.protection.copy()
                header_cell_2.alignment = ref_header_cell.alignment.copy()
        
        # 填充第3到第25行的数据（从CCE列第3行到CCF列第25行复制数据）并保持格式
        import random
        for day_idx in range(days_count):
            for row in range(3, 26):  # 第3行到第25行
                # 每天的数据列起始索引
                day_start_col = ccg_column_index + 1 + (day_idx * 2)
                
                # 获取参考单元格用于复制格式
                ref_data_cell = ws_data_inspection.cell(row=row, column=ccg_column_index)
                
                # 从CCE列复制数据到工单数量列（CCE是第2111列）
                source_cell_1 = ws_data_inspection.cell(row=row, column=2111)  # CCE列
                # 添加-5%到+5%的浮动并转换为整数
                original_value_1 = source_cell_1.value
                if isinstance(original_value_1, (int, float)) and original_value_1 != 0:
                    # 计算浮动后的值
                    fluctuation_1 = random.uniform(-0.05, 0.05)
                    fluctuated_value_1 = int(round(original_value_1 * (1 + fluctuation_1)))
                    data_cell_1 = ws_data_inspection.cell(row=row, column=day_start_col, value=fluctuated_value_1)
                else:
                    data_cell_1 = ws_data_inspection.cell(row=row, column=day_start_col, value=original_value_1)
                    
                # 复制格式
                if ref_data_cell.has_style:
                    data_cell_1.font = ref_data_cell.font.copy()
                    data_cell_1.border = ref_data_cell.border.copy()
                    data_cell_1.fill = ref_data_cell.fill.copy()
                    data_cell_1.number_format = ref_data_cell.number_format
                    data_cell_1.protection = ref_data_cell.protection.copy()
                    data_cell_1.alignment = ref_data_cell.alignment.copy()
                
                # 从CCF列复制数据到平均时间列（CCF是第2112列）
                source_cell_2 = ws_data_inspection.cell(row=row, column=2112)  # CCF列
                # 添加-5%到+5%的浮动并保留两位小数
                original_value_2 = source_cell_2.value
                if isinstance(original_value_2, (int, float)) and original_value_2 != 0:
                    # 计算浮动后的值
                    fluctuation_2 = random.uniform(-0.05, 0.05)
                    fluctuated_value_2 = round(original_value_2 * (1 + fluctuation_2), 2)
                    data_cell_2 = ws_data_inspection.cell(row=row, column=day_start_col + 1, value=fluctuated_value_2)
                else:
                    data_cell_2 = ws_data_inspection.cell(row=row, column=day_start_col + 1, value=original_value_2)

                # 复制格式
                if ref_data_cell.has_style:
                    data_cell_2.font = ref_data_cell.font.copy()
                    data_cell_2.border = ref_data_cell.border.copy()
                    data_cell_2.fill = ref_data_cell.fill.copy()
                    data_cell_2.number_format = ref_data_cell.number_format
                    data_cell_2.protection = ref_data_cell.protection.copy()
                    data_cell_2.alignment = ref_data_cell.alignment.copy()

    # 保存新文件
    output_file = os.path.join(output_dir, f'report_{date_str}.xlsx')
    wb.save(output_file)
    
    print(f'Saved report for {current_date.strftime("%Y-%m-%d")}')
    current_date += timedelta(days=1)

print("All reports generated successfully.")