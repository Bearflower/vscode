import os
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook

def generate_inspection_reports(template_path: str, output_dir: str, start_date: datetime, num_days: int):
    """
    根据给定的模板文件批量生成每日巡检报告
    
    Args:
        template_path (str): 模板文件路径
        output_dir (str): 输出目录路径
        start_date (datetime): 报告开始日期
        num_days (int): 需要生成的报告天数
    """
    
    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    # 加载模板文件
    try:
        template_wb = load_workbook(template_path)
    except Exception as e:
        print(f"无法加载模板文件 {template_path}: {e}")
        return
        
    # 依次为每一天生成报告
    for i in range(num_days):
        current_date = start_date + timedelta(days=i)
        formatted_date = current_date.strftime("%Y%m%d")
        
        # 创建新工作簿并复制模板内容
        new_wb = Workbook()
        new_wb.remove(new_wb.active)  # 删除默认创建的工作表
        
        # 将模板中的每个工作表复制到新的工作簿中
        for sheet_name in template_wb.sheetnames:
            source_sheet = template_wb[sheet_name]
            
            # 创建同名的新工作表
            target_sheet = new_wb.create_sheet(title=sheet_name)
            
            # 复制单元格数据和样式（简化版）
            for row in source_sheet.iter_rows(values_only=False):
                for cell in row:
                    target_cell = target_sheet[cell.coordinate]
                    target_cell.value = cell.value
                    
                    # 如果需要保留更多样式信息，可在此处添加
                    
        # 定义输出文件名
        output_file = os.path.join(output_dir, f"穗碳巡检{formatted_date}.xlsx")
        
        # 保存新生成的Excel文件
        try:
            new_wb.save(output_file)
            print(f"已生成巡检报告: {output_file}")
        except Exception as e:
            print(f"保存文件失败 {output_file}: {e}")

if __name__ == "__main__":
    # 设置参数
    template_file = "/Users/yl/vscode/xunjianreport/report_demo/st20251213.xlsx"
    reports_output_dir = "/Users/yl/vscode/xunjianreport/stdaily_reports"
    start_from_date = datetime(2025, 7, 1)
    total_days_to_generate = 10
    
    # 执行生成操作
    generate_inspection_reports(
        template_path=template_file,
        output_dir=reports_output_dir,
        start_date=start_from_date,
        num_days=total_days_to_generate
    )